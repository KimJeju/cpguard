"""고객 제출용 '분석목록표' 엑셀 내보내기.

현장 진단 산출물 규격(preprocess_scan_results.py / insert_column.py 에서 옮겨옴):
  - 14개 고정 컬럼: ID·위험도·언어·체커명·라인·파일명·점검 대상·함수·경로·체커 설명·
    이슈 의견·조치 여부·조치 방법·소스 코드
  - 조치 여부는 N 으로 초기화, 조치 방법은 비워 두어 담당자가 채운다
  - 위험도별 셀 음영, 회색 헤더, 첫 행 고정, 자동 필터
  - 요약 시트: 위험도/규칙/파일별 집계
탐지값이 마스킹된 채로 들어가므로 산출물 자체가 유출원이 되지 않는다.
"""
from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path

from .finding import Finding

COLUMNS = [
    "ID", "위험도", "언어", "체커명", "라인", "파일명", "점검 대상(Y/N)", "함수", "경로",
    "체커 설명", "이슈 의견", "조치 여부(Y/N)", "조치 방법", "소스 코드",
]
COLUMN_WIDTHS = {
    "A": 7.5, "B": 9, "C": 10, "D": 26, "E": 7.5, "F": 24, "G": 13, "H": 18,
    "I": 44, "J": 40, "K": 44, "L": 13, "M": 30, "N": 70,
}
SEVERITY_KO = {"critical": "심각", "high": "높음", "medium": "중간", "low": "낮음", "info": "정보"}
SEVERITY_FILL = {
    "critical": "F8B4B4", "high": "F8CBAD", "medium": "FFE699", "low": "D9E1F2", "info": "EDEDED",
}
LANG_BY_EXT = {
    ".js": "JavaScript", ".jsx": "JavaScript", ".mjs": "JavaScript", ".cjs": "JavaScript",
    ".ts": "TypeScript", ".tsx": "TypeScript", ".php": "PHP", ".phtml": "PHP", ".inc": "PHP",
    ".py": "Python", ".sql": "SQL", ".xml": "XML", ".yml": "YAML", ".yaml": "YAML",
    ".json": "JSON", ".properties": "Properties", ".env": "Config", ".ini": "Config",
    ".java": "Java", ".kt": "Kotlin", ".kts": "Kotlin", ".cs": "C#", ".go": "Go",
    ".rb": "Ruby", ".rake": "Ruby", ".swift": "Swift",
    ".cpp": "C++", ".cc": "C++", ".cxx": "C++", ".hpp": "C++", ".hh": "C++", ".hxx": "C++",
    ".c": "C", ".h": "C/C++",
}

# 규칙 접두 → 표준 이슈 의견. 담당자에게 '왜'와 '어떻게'를 같은 문장으로 준다.
REMEDIATION = {
    "sqli": "사용자 입력이 SQL 문에 결합됩니다. 파라미터 바인딩(Prepared Statement)으로 전환하고 입력값을 화이트리스트로 검증하세요.",
    "command-injection": "사용자 입력이 시스템 명령에 결합됩니다. 셸을 거치지 않는 API(인자 배열)를 쓰고 escapeshellarg/shlex.quote 등으로 인자를 감싸세요.",
    "code-injection": "사용자 입력이 eval 등 코드 실행 함수에 전달됩니다. 동적 코드 실행을 제거하고 필요하면 안전한 파서나 화이트리스트 디스패치로 대체하세요.",
    "xss": "사용자 입력이 정제 없이 출력됩니다. 출력 문맥에 맞게 인코딩(htmlspecialchars/escape)하고 React 는 dangerouslySetInnerHTML 사용을 피하세요.",
    "path-traversal": "사용자 입력이 파일 경로로 쓰입니다. basename 으로 정규화하고 허용 디렉터리 밖 접근을 차단하세요.",
    "file-inclusion": "사용자 입력이 include/require 에 쓰입니다. 포함 가능한 파일을 고정 목록으로 제한하세요.",
    "ssrf": "사용자 입력이 서버측 요청 대상이 됩니다. 허용 호스트 목록으로 검증하고 내부망 주소를 차단하세요.",
    "open-redirect": "사용자 입력이 리다이렉트 대상이 됩니다. 상대 경로만 허용하거나 허용 도메인 목록으로 검증하세요.",
    "secret.": "비밀정보를 소스에서 제거하고 환경변수·비밀 관리 서비스로 옮기세요. 이미 커밋된 값은 폐기·재발급하세요.",
    "vendor.": "벤더 API 키가 노출되었습니다. 즉시 재발급하고 소스에서 제거한 뒤 환경변수로 주입하세요. 저장소 이력도 정리하세요.",
    "pii.": "개인정보가 소스/덤프에 평문으로 존재합니다. 즉시 삭제하고 필요 시 암호화·마스킹 저장으로 전환하세요. 개인정보보호법상 안전조치 의무 대상입니다.",
    "web.tls": "TLS 인증서 검증을 켜세요. 자체 서명 인증서는 신뢰 저장소에 등록해 해결합니다.",
    "web.token-in-web-storage": "토큰을 웹 스토리지에 두면 XSS 시 탈취됩니다. HttpOnly·Secure 쿠키로 옮기세요.",
    "web.": "정제되지 않은 값이 DOM 에 직접 들어갑니다. DOMPurify 등으로 정제하거나 textContent 를 사용하세요.",
    "crypto.": "취약한 알고리즘입니다. SHA-256 이상, AES-GCM, 암호학적 난수(crypto.randomBytes/secrets)로 교체하세요.",
    "hygiene.debug": "운영 코드에서 디버그 코드·디버그 모드를 제거하세요. 정보 노출과 성능 문제의 원인이 됩니다.",
    "hygiene.": "운영 반영 전 정리 대상입니다.",
    "infra.default-test-account": "기본/테스트 계정을 제거하고 초기 비밀번호를 강제 변경하세요.",
    "infra.": "내부 인프라 정보를 소스에서 제거하고 설정으로 외부화하세요.",
}


# ── 영문(외국 대상) — lang="en" 일 때 사용 ──────────────────────────────
COLUMNS_EN = [
    "ID", "Severity", "Language", "Checker", "Line", "File", "In scope(Y/N)", "Function", "Path",
    "Checker description", "Issue note", "Fixed(Y/N)", "Remediation", "Source code",
]
SEVERITY_EN = {"critical": "Critical", "high": "High", "medium": "Medium", "low": "Low", "info": "Info"}
REMEDIATION_EN_XLSX = {
    "sqli": "User input is concatenated into SQL. Switch to parameter binding (prepared statements) and validate input against an allowlist.",
    "command-injection": "User input is concatenated into a system command. Use a non-shell API (argument array) and wrap arguments with escapeshellarg/shlex.quote.",
    "code-injection": "User input reaches a code-execution function such as eval. Remove dynamic code execution; if needed replace with a safe parser or allowlist dispatch.",
    "xss": "User input is output without sanitization. Encode for the output context (htmlspecialchars/escape) and avoid dangerouslySetInnerHTML in React.",
    "path-traversal": "User input is used as a file path. Normalize with basename and block access outside the allowed directory.",
    "file-inclusion": "User input is used in include/require. Restrict includable files to a fixed list.",
    "ssrf": "User input becomes a server-side request target. Validate against an allowed-host list and block internal addresses.",
    "open-redirect": "User input becomes a redirect target. Allow only relative paths or validate against an allowed-domain list.",
    "secret.": "Remove the secret from the source and move it to environment variables / a secret manager. Revoke and rotate any already-committed value.",
    "vendor.": "A vendor API key is exposed. Rotate it immediately, remove it from source, inject via environment variables, and clean the repository history.",
    "pii.": "Personal data exists in plaintext in the source/dump. Delete it immediately and switch to encrypted/masked storage as needed; this is subject to statutory safeguards under privacy law.",
    "web.tls": "Enable TLS certificate verification. Register self-signed certificates in the trust store.",
    "web.token-in-web-storage": "Tokens in web storage are stealable via XSS. Move them to HttpOnly/Secure cookies.",
    "web.": "Unsanitized values reach the DOM directly. Sanitize with DOMPurify or use textContent.",
    "crypto.": "A weak algorithm is used. Replace with SHA-256+, AES-GCM, and cryptographic RNG (crypto.randomBytes/secrets).",
    "hygiene.debug": "Remove debug code / debug mode from production code — it causes information disclosure and performance issues.",
    "hygiene.": "Clean up before production.",
    "infra.default-test-account": "Remove default/test accounts and force an initial password change.",
    "infra.": "Remove internal infrastructure details from source and externalize them to configuration.",
}
# 요약/시트 라벨
LABELS_EN = {
    "요약": "Summary", "분석목록표": "Analysis Sheet",
    "소스코드 취약점 진단 분석목록표": "Source Code Vulnerability Analysis Sheet",
    "프로젝트": "Project", "생성 일시": "Generated", "총 탐지 건수": "Total findings",
    "위험도": "Severity", "건수": "Count", "규칙": "Rule", "파일 (상위 30)": "File (top 30)",
}


def _remediation(rule_id: str, lang: str = "ko") -> str:
    table = REMEDIATION_EN_XLSX if lang == "en" else REMEDIATION
    for key, text in table.items():
        if key in rule_id:
            return text
    return ""


def _language_of(path: str) -> str:
    return LANG_BY_EXT.get(Path(path).suffix.lower(), Path(path).suffix.lstrip(".").upper() or "-")


def _source_text(f: Finding) -> str:
    """흐름형은 단계 전체를, 패턴형은 매칭 줄(마스킹)을 넣는다."""
    if f.category == "flow" or len(f.steps) > 1:
        return "\n".join(f"[{s.kind}] L{s.loc.start_line}: {s.code}" for s in f.steps)
    return f.steps[0].code if f.steps else ""


def _rel(path: str, base: Path | None) -> str:
    from ..extract import strip_longpath
    path = strip_longpath(path)          # Windows 긴 경로 탐색용 \\?\ 접두 제거
    if base is None:
        return path
    try:
        return str(Path(path).resolve().relative_to(Path(strip_longpath(base)))).replace("\\", "/")
    except Exception:
        return Path(path).name


def to_rows(findings: list[Finding], base: str | Path | None = None,
            audit: dict[str, str] | None = None, lang: str = "ko") -> list[list]:
    from ..i18n import tr
    en = lang == "en"
    sevmap = SEVERITY_EN if en else SEVERITY_KO
    base_path = Path(base).resolve() if base else None
    audit = audit or {}
    rows = []
    for i, f in enumerate(findings):
        rel = _rel(f.sink.loc.file, base_path)
        opinion = tr(f.message, lang)
        if f.fp_hint:
            opinion += (" (env/example signal on the same line — possible false positive, verify)"
                        if en else " (같은 줄에 환경변수/예시 신호가 있어 오탐 가능성 있음 — 확인 필요)")
        if f.verdict:
            opinion += (f" [LLM verdict: {f.verdict}]" if en else f" [LLM 판정: {f.verdict}]")
        done = "Y" if audit.get(str(i)) in ("fixed",) else "N"
        rows.append([
            i + 1,
            sevmap.get(f.severity, f.severity),
            _language_of(rel),
            f.rule_id,
            f.sink.loc.start_line or "",
            Path(rel).name,
            "Y",
            "",
            rel,
            f"{f.cwe}{' · ' + f.owasp if f.owasp else ''}",
            opinion,
            done,
            _remediation(f.rule_id, lang),
            _source_text(f),
        ])
    return rows


def write_workbook(findings: list[Finding], out_path: str | Path,
                   project: str = "결과", base: str | Path | None = None,
                   audit: dict[str, str] | None = None, lang: str = "ko") -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    en = lang == "en"
    L = (lambda s: LABELS_EN.get(s, s)) if en else (lambda s: s)
    sevmap = SEVERITY_EN if en else SEVERITY_KO
    cols = COLUMNS_EN if en else COLUMNS

    wb = Workbook()

    # ---- 요약 ----
    s = wb.active
    s.title = L("요약")
    bold = Font(name="맑은 고딕", size=11, bold=True)
    s.append([L("소스코드 취약점 진단 분석목록표")]); s["A1"].font = Font(name="맑은 고딕", size=14, bold=True)
    s.append([])
    s.append([L("프로젝트"), project])
    s.append([L("생성 일시"), datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    s.append([L("총 탐지 건수"), len(findings)])
    s.append([])
    s.append([L("위험도"), L("건수")]); s.cell(s.max_row, 1).font = bold; s.cell(s.max_row, 2).font = bold
    for sev in ("critical", "high", "medium", "low", "info"):
        n = sum(1 for f in findings if f.severity == sev)
        if n:
            s.append([sevmap[sev], n])
    s.append([])
    s.append([L("규칙"), L("건수")]); s.cell(s.max_row, 1).font = bold; s.cell(s.max_row, 2).font = bold
    for rid, n in Counter(f.rule_id for f in findings).most_common():
        s.append([rid, n])
    s.append([])
    s.append([L("파일 (상위 30)"), L("건수")]); s.cell(s.max_row, 1).font = bold; s.cell(s.max_row, 2).font = bold
    base_path = Path(base).resolve() if base else None
    for path, n in Counter(_rel(f.sink.loc.file, base_path) for f in findings).most_common(30):
        s.append([path, n])
    s.column_dimensions["A"].width = 48
    s.column_dimensions["B"].width = 12

    # ---- 분석목록표 ----
    ws = wb.create_sheet(L("분석목록표"))
    thin = Side(style="thin", color="FF000000")
    thick = Side(style="thick", color="FF000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thick)
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="맑은 고딕", size=11, bold=True)
    data_font = Font(name="맑은 고딕", size=10)
    header_fill = PatternFill(fill_type="solid", fgColor="FFBFBFBF")
    header_align = Alignment(horizontal="center", vertical="center")
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="top")

    ws.append(cols)
    for c in ws[1]:
        c.font, c.fill, c.border, c.alignment = header_font, header_fill, header_border, header_align
    ws.row_dimensions[1].height = 20

    rows = to_rows(findings, base, audit, lang)
    for r, f in zip(rows, findings):
        ws.append(r)
        rn = ws.max_row
        ws.row_dimensions[rn].height = min(15 * max(2, str(r[13]).count("\n") + 2), 120)
        for c in ws[rn]:
            c.font, c.border, c.alignment = data_font, data_border, data_align
        for col in (1, 2, 3, 5, 7, 12):
            ws.cell(rn, col).alignment = center
        fill = SEVERITY_FILL.get(f.severity)
        if fill:
            ws.cell(rn, 2).fill = PatternFill(fill_type="solid", fgColor=fill)

    for col, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(ws.max_row, 1)}"

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
