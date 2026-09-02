"""사용자 현장 스크립트에서 옮겨온 기능 검증 — 검증기·마스킹·벤더키·오탐힌트·위험파일명·엑셀.

주의: 실제 개인정보 형식의 문자열은 체크섬을 통과하는 합성값만 쓴다.
"""
import pytest

from cpguard import patterns
from cpguard.patterns import load_pattern_rules, scan_filename, scan_text

RULES = load_pattern_rules()


def hits(src: str, file: str = "dump.sql", ext_rules=None):
    rules = ext_rules if ext_rules is not None else patterns.rules_for(RULES, None, "." + file.rsplit(".", 1)[-1])
    return {f.rule_id: f for f in scan_text(src, file, rules)}


# ---------- 검증기: 정규식이 잡은 후보를 체크섬으로 확정 ----------

def test_rrn_checksum_validator():
    assert patterns._rrn_ok("900101-1234568")       # 계산상 유효
    assert not patterns._rrn_ok("900101-1234567")   # 체크섬 불일치
    assert not patterns._rrn_ok("901301-1234568")   # 13월


def test_luhn_validator():
    assert patterns._luhn_ok("4111 1111 1111 1111")
    assert not patterns._luhn_ok("4111 1111 1111 1112")


def test_rrn_detected_only_when_checksum_passes():
    ok = hits("INSERT INTO m VALUES ('900101-1234568');")
    bad = hits("INSERT INTO m VALUES ('900101-1234567');")
    assert "pii.kr-resident-number" in ok
    assert "pii.kr-resident-number" not in bad


def test_card_detected_only_when_luhn_passes():
    assert "pii.credit-card-number" in hits("card: 4111 1111 1111 1111")
    assert "pii.credit-card-number" not in hits("card: 4111 1111 1111 1112")


# ---------- 마스킹: 산출물이 유출원이 되지 않게 ----------

def test_rrn_is_masked_in_finding():
    f = hits("VALUES ('900101-1234568')")["pii.kr-resident-number"]
    assert f.matched_value == "900101-1******"
    assert "1234568" not in f.steps[0].code          # 스니펫에도 원본이 없어야 한다


def test_card_and_phone_and_email_masks():
    assert patterns._mask_card("4111111111111111") == "4111-****-****-1111"
    assert patterns._mask_phone("010-1234-5678") == "010-****-5678"
    assert patterns._mask_email("kimjeju@example.com") == "ki***@example.com"


def test_secret_value_is_masked():
    f = hits('password = "SuperSecret99!"', "app.properties")["secret.config-file-credential"]
    assert "SuperSecret99!" not in f.steps[0].code
    assert f.matched_value.startswith("Supe") and "*" in f.matched_value


# ---------- 벤더 키 (형식 자체가 식별자 → 신뢰도 high) ----------

# 벤더 키 형식 문자열을 소스에 그대로 적으면 GitHub 푸시 보호가 시크릿으로 보고 막는다
# (실제로 겪음). 테스트 페이로드는 조각을 이어붙여 만든다.
def _k(prefix: str, body: str) -> str:
    return prefix + body


@pytest.mark.parametrize("rule_id,src", [
    ("vendor.aws-access-key-id", _k("AKIA", "IOSFODNN7QWERTYU")),
    ("vendor.google-api-key", _k("AIza", "SyA1234567890abcdefghijklmnopqrstu")),
    ("vendor.github-token", _k("ghp_", "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789")),
    ("vendor.slack-token", _k("xoxb-", "123456789012-abcdefghijklmnop")),
    ("vendor.stripe-key", _k("sk_live_", "abcdefghijklmnop")),
    ("vendor.jwt-token", _k("eyJhbGciOiJIUzI1NiJ9.", "eyJzdWIiOiIxMjM0NTY3ODkwIn0.dozjgNryP4J3jVmNHl0w5N_XgL0n3I9PlFUP0THsR8U")),
])
def test_vendor_keys(rule_id, src):
    f = hits(f"x = '{src}'", "config.txt")
    assert rule_id in f
    # 형식이 고유한 키는 high, JWT 처럼 형식만으로는 애매한 것은 medium 으로 둔다
    assert f[rule_id].precision in ("high", "medium")
    assert f[rule_id].category == "secret"


def test_aws_example_key_excluded():
    assert "vendor.aws-access-key-id" not in hits("AKIAIOSFODNN7EXAMPLE", "a.txt")


# ---------- 오탐 힌트: 제외하지 않고 신뢰도만 낮춘다 ----------

def test_fp_hint_lowers_confidence_but_keeps_finding():
    f = hits(f"token = '{_k('xoxb-', '123456789012-abcdefghijklmnop')}'  # placeholder", "a.txt")
    assert "vendor.slack-token" in f
    assert f["vendor.slack-token"].fp_hint is True
    assert f["vendor.slack-token"].precision == "medium"   # high → medium


# ---------- 설정파일 전용 규칙은 소스 코드에 적용되지 않는다 ----------

def test_config_credential_rule_skips_source_code():
    """user.password = hash_password() 같은 코드를 잡으면 안 된다(실전 프로젝트 오탐)."""
    src_rules = patterns.rules_for(RULES, "python", ".py")
    assert "secret.config-file-credential" not in {r.id for r in src_rules}
    cfg_rules = patterns.rules_for(RULES, None, ".properties")
    assert "secret.config-file-credential" in {r.id for r in cfg_rules}


def test_column_hint_rule_only_on_dumps():
    assert "pii.sensitive-column-name" in hits("CREATE TABLE u (jumin VARCHAR(14));", "schema.sql")
    py_rules = patterns.rules_for(RULES, "python", ".py")
    assert "pii.sensitive-column-name" not in {r.id for r in py_rules}


# ---------- 위험 파일명 ----------

@pytest.mark.parametrize("name", [".env", ".env.production", "id_rsa", "server.pem",
                                  "keystore.jks", "backup.bak", "wp-config.php"])
def test_risky_filenames(name):
    f = scan_filename(name)
    assert f is not None and f.rule_id == "secret.risky-filename"


def test_normal_filename_is_fine():
    assert scan_filename("app.py") is None
    assert scan_filename("index.tsx") is None


# ---------- 엑셀 분석목록표 ----------

def test_excel_export(tmp_path):
    from openpyxl import load_workbook
    from cpguard.report import excel

    fs = list(hits("VALUES ('900101-1234568')").values())
    fs += list(hits("AKIAIOSFODNN7QWERTYU", "conf.txt").values())
    out = excel.write_workbook(fs, tmp_path / "r.xlsx", project="샘플")

    wb = load_workbook(out)
    assert wb.sheetnames == ["요약", "분석목록표"]
    ws = wb["분석목록표"]
    assert [c.value for c in ws[1]] == excel.COLUMNS
    assert ws.max_row == 1 + len(fs)
    row = [c.value for c in ws[2]]
    assert row[1] in ("심각", "높음", "중간", "낮음", "정보")   # 위험도 한글
    assert row[6] == "Y" and row[11] == "N"                    # 점검대상 Y, 조치여부 N
    assert row[12]                                              # 조치 방법(표준 의견) 채워짐
    assert "1234568" not in str(row[13])                        # 소스코드 열도 마스킹
